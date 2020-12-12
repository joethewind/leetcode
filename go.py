# -*- coding: utf-8 -*-

import os,sys,openpyxl,re
from openpyxl.styles import PatternFill
from openpyxl.styles import Border,Side,Alignment
import tempfile,time
#from utils import *

#子载波间隔与时隙对应表
BWP_slot_dict = {'kHz15':1,'kHz30':2,'kHz60':4,'kHz120':8,'kHz240':16}
#ncs计算方式
kct2 = {'0':4,'4':'0','2':6,'6':2,'1':5,'5':1,'3':7,'7':3}
kct4 = {'0':6,'3':9,'6':0,'9':3,'1':7,'4':10,'7':1,'10':4,'2':8,'5':11,'8':2,'11':5}

'''
#获取excel路径、log路径等
'''
def init():
    exl_path = ''
    log_path = ''
    path = sys.path[0]
    for root, dirs, files in os.walk(path):
        if len(files) > 0:
            for file in files:
                if '.log' in file:
                    log_path = os.path.join(root, file)
    file = r'\resources\多用户资源分配日志提取及分析方法.xlsm'
    exl_path = path + file
    if exl_path == '' or log_path == '':
        raise ('请检查日志或Excel文件是否存在')
    else:
        read_workbook = openpyxl.load_workbook(exl_path,keep_vba=True)
        fmt = '%Y%m%d%H%M%S'
        Date = time.strftime(fmt,time.localtime(time.time()))
        excel_path = exl_path.replace(r'\resources','').replace('.xlsm','') + '_' + Date + '.xlsm'
        read_workbook.save(excel_path)
        return excel_path,log_path

'''
对log文件进行解析
'''
def get_data(path):
    #临时储存数据
    data_list = []
    #释放msg的list
    valid_data = []
    #接入msg的list
    swich_data = []
    #ue冲突用接入list
    #ue冲突用终端去注册list
    request_transfer = []
    #invalid无效广播小区级集合
    invalid_data = []
    #自动录入参数ue级集合
    auto_data = []

    row_start = 0
    row_stop = 0
    #用于标识接入msg
    tag = False
    with open(path,encoding='utf-8') as f:
        lines = f.readlines()
        for i in range(len(lines)):
            data = ''
            #如果包含PCO_RRC_NR_DL_DCCH_Message，message c1 : rrcReconfiguration :则开始向临时文件中存
            #如果包含NR-DL-DCCH-Message则停止加入
            if 'PCO_RRC_NR_DL_DCCH_Message' in lines[i]:
                row_start = i
            if 'NR-DL-DCCH-Message' in lines[i]:
                row_stop = i
                data = get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,lines)
                data_list.append(data)

            #用于UE冲突分析
            if 'PCO_CFG_NAS_UL_NMM_Message' in lines[i]:
                row_start = i
            if 'NMM-UL-Message' in lines[i]:
                row_stop = i
                data = get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,lines)
                data_list.append(data)
            #用于UE冲突分析,接入消息
            if 'PCO_RRC_NR_UL_CCCH_Message' in lines[i]:
                row_start = i
            if 'NR-UL-CCCH-Message' in lines[i]:
                row_stop = i
                data = get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,lines)
                data_list.append(data)
            #去注册消息
            if 'PCO_RRC_NR_UL_DCCH_Message' in lines[i]:
                row_start = i
            if 'NR-UL-DCCH-Message' in lines[i]:
                row_stop = i
                data = get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,lines)
                data_list.append(data)
                
            #无效消息，获取广播级参数
            if 'PCO_RRC_NR_BCCH_DL_SCH_Message' in lines[i]:
                row_start = i
            if 'NR-BCCH-DL-SCH-Message' in lines[i]:
                row_stop = i
                data = get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,lines)
                data_list.append(data)
            #pucch相关参数rrcsetup,获取UE级参数
            if 'PCO_RRC_NR_DL_CCCH_Message' in lines[i]:
                row_start = i
            if 'NR-DL-CCCH-Message' in lines[i]:
                row_stop = i
                data = get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,lines)
                data_list.append(data)

    for data in data_list:
        if 'message c1 : rrcReconfiguration :' in data:
            valid_data.append(data)
            request_transfer.append('重配置：')
            request_transfer.append(data)
        #ue接入次数
        if 'nrDeregistrationRequestUeOriginating' in data:
            tag = True
        if ('switchOff switchOff-SwitchOff' in data) and (tag == True):
            swich_data.append(data)
        #接入消息
        if 'message c1 : rrcSetupRequest :' in data:
            request_transfer.append('接入：')
            request_transfer.append(data)
        #去注册消息
        if 'message c1 : ulInformationTransfer :' in data:
            request_transfer.append('去注册：')
            request_transfer.append(data)
        #广播消息
        if 'message c1 : systemInformationBlockType1 :' in data:
            invalid_data.append(data)
        if 'message c1 : rrcSetup :' in data:
            auto_data.append(data)
    return valid_data,swich_data,request_transfer,invalid_data,auto_data

def get_PCO_RRC_NR_DL_DCCH_Message(row_start,row_stop,list):
    msg = ''
    for i in range(row_start,row_stop):
        msg = msg + list[i]
    return msg
    
def get_PCO_RRC_NR_DL_CCCH_Message(row_start,row_stop,list):
    msg = ''
    for i in range(row_start,row_stop):
        if msg != '':
            msg = msg + '\n' + list[i]
        else:
            msg = list[i] + '\n'
    return msg

'''
获取data中的Ue Index、Time、gapOffset(ms)、mgl(ms)、mgrp(ms)、mgta、srs-ResourceSetId、usage、
    srs-ResourceId、nrofSRS-Ports、transmissionComb、combOffset-n2、cyclicShift-n2、端口2占用ncs、
    startPosition、nrofSymbols、repetitionFactor、freqDomainPosition、freqDomainShift、c-SRS、
    b-SRS、b-hop、periodicityAndOffset-p、Offset、srs落入GAP冲突比例、UE级别SRS与GAP冲突概率
    各UE资源是否冲突、单子帧包含的时隙个数，并记录至xml中
'''
def set_GAP_AND_SRS(path,data_list,swich_data_list):
    #打开excel
    read_workbook = openpyxl.load_workbook(path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']
    #read_SRS_GAP_sheet = read_workbook.active
    #清空原有数据
    #excel清空填充颜色
    white = PatternFill("solid",fgColor="ffffff")
    #excel加框
    thin = Side(border_style="thin",color="000000")
    nrows_max = read_SRS_GAP_sheet.max_row
    ncols_max = read_SRS_GAP_sheet.max_column
    #删除前判断是否有合并单元格，有则拆分单元格
    mt = read_SRS_GAP_sheet.merged_cells
    cr = []
    for ma in mt :
        row1,row2,col1,col2 = ma.min_row,ma.max_row,ma.min_col,ma.max_col
        if row2 - row1 > 0:
            cr.append((row1,row2,col1,col2))
    for r in cr:
        read_SRS_GAP_sheet.unmerge_cells(start_row=r[0], start_column=r[2], end_row=r[1], end_column=r[3])
    for i in range(2,nrows_max+1):
        for j in range(1,ncols_max+1):
            #背景颜色为白
            read_SRS_GAP_sheet.cell(i,j,'').fill=white
            #添加边框为黑色
            read_SRS_GAP_sheet.cell(i,j,'').border=Border(top = thin,right = thin,bottom = thin,left = thin)
            #数据居中
            read_SRS_GAP_sheet.cell(i,j,'').alignment=Alignment(horizontal='center',vertical='center')
    
    read_workbook.save(path)  
    print('正在生成excel中...')
    #获取SRS&GAP原始数据sheet页首行名字当动态列名
    #获取数据行数，用于数据追加
    read_workbook = openpyxl.load_workbook(path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']
    nrows = read_SRS_GAP_sheet.min_row
    cols = read_SRS_GAP_sheet.max_column
    value = []
    for i in range(1,cols + 1):
        cell_value = read_SRS_GAP_sheet.cell(row=1,column=i).value
        value.append(cell_value)
    #匹配列名
    if 'Ue Index:' in value:
        ue_index_col = value.index('Ue Index:')
    if 'Time' in value:
        time_col = value.index('Time')
    if 'BWP编号' in value:
        bwp_id_col = value.index('BWP编号')
    if 'gapOffset(ms)' in value:
        gapOffset_ms_col = value.index('gapOffset(ms)')
    if 'mgl(ms)' in value:
        mgl_ms_col = value.index('mgl(ms)')
    if 'mgrp(ms)' in value:
        mgrp_ms_col = value.index('mgrp(ms)')
    if 'mgta' in value:
        mgta_col = value.index('mgta')
    if 'srs-ResourceSetId' in value:
        srs_ResourceSetId_col = value.index('srs-ResourceSetId')
    if 'usage' in value:
        usage_col = value.index('usage')
    if 'srs-ResourceId' in value:
        srs_ResourceId_col = value.index('srs-ResourceId')
    if 'nrofSRS-Ports' in value:
        nrofSRS_Ports_col = value.index('nrofSRS-Ports')
    if 'transmissionComb' in value:
        transmissionComb_col = value.index('transmissionComb')
    if 'combOffset-n2' in value:
        combOffset_n2_col = value.index('combOffset-n2')
    if 'cyclicShift-n2' in value:
        cyclicShift_n2_col = value.index('cyclicShift-n2')
    if '端口2占用ncs' in value:
        port2_ncs_col = value.index('端口2占用ncs')
    if 'startPosition' in value:
        startPosition_col = value.index('startPosition')
    if 'nrofSymbols' in value:
        nrofSymbols_col = value.index('nrofSymbols')
    if 'repetitionFactor' in value:
        repetitionFactor_col = value.index('repetitionFactor')
    if 'freqDomainPosition' in value:
        freqDomainPosition_col = value.index('freqDomainPosition')
    if 'freqDomainShift' in value:
        freqDomainShift_col = value.index('freqDomainShift')
    if 'c-SRS' in value:
        c_SRS_col = value.index('c-SRS')
    if 'b-SRS' in value:
        b_SRS_col = value.index('b-SRS')
    if 'b-hop' in value:
        b_hop_col = value.index('b-hop')
    if 'periodicityAndOffset-p' in value:
        periodicityAndOffset_p_col = value.index('periodicityAndOffset-p')
    if 'Offset' in value:
        Offset_col = value.index('Offset')
    if 'srs落入GAP冲突比例' in value:
        srs_GAP_col = value.index('srs落入GAP冲突比例')
    if 'UE级别SRS与GAP冲突概率' in value:
        ue_srs_gap_col = value.index('UE级别SRS与GAP冲突概率')
    if '各UE资源是否冲突' in value:
        ue_resource_col = value.index('各UE资源是否冲突')
    if '单子帧包含的时隙个数' in value:
        slot_num_col = value.index('单子帧包含的时隙个数')
    if 'UE第几次接入' in value:
        ue_swich_col = value.index('UE第几次接入')
    
    #383个接入数据
    #用于储存时间戳及ue_index的list，dict有冲突key，弃用
    swich = []
    for swich_data in swich_data_list:
        list = []
        #对每一个数据去除特殊字符，并按照回车切分
        datas = swich_data.replace('\t','').replace('{','').replace('}','').strip(' ')
        for data in datas.split('\n'):
            value = data.strip(' ')
            if value != '':
                list.append(value)
        #UE Index在第一行中
        first_row_list = list[0].split(',')
        for ret in first_row_list:
            #ue_index
            if 'Index' in ret:
                ue_index_value = ret.split(':')[1]          
                swich.append(ue_index_value)
        #time
        time_value = first_row_list[0].split(' ')[0]
        if 12 == len(time_value):
            swich.append(time_value)
            
    #877个有效释放数据
    #rrcReconfiguration用于储存time及ue_index，用于判断ue接入次数
    rrcReconfiguration = []
    for valid_data in data_list:
        list = []
        #对每一个数据去除特殊字符，并按照回车切分
        datas = valid_data.replace('\t','').replace('{','').replace('}','').strip(' ')
        #对每行数据进行去除前后的空格
        for data in datas.split('\n'):
            value = data.strip(' ')
            if value != '':                
                list.append(value)
        
        #判断消息长度，是否有必要存入excel中
        if len(list) < 400:
            continue

        #BWP编号
        if 'spCellConfig' in list:            
            spCellConfig_index = list.index('spCellConfig')
            if 'spCellConfigDedicated' == list[spCellConfig_index+1]:
                if 'downlinkBWP-ToAddModList' == list[spCellConfig_index+2]:
                    BWP_id = list[spCellConfig_index+3]
                    BWP_id_value = BWP_id.split(' ')[1].replace(',','')
                    read_SRS_GAP_sheet.cell(nrows+1,bwp_id_col+1,BWP_id_value) 
        
                #单子帧包含的时隙个数
                if 'subcarrierSpacing' in list[spCellConfig_index+7]:
                    slot_value = list[spCellConfig_index+7].split(' ')[1]
                    for i in BWP_slot_dict.keys():
                        if i == slot_value:
                            subcarrierSpacing_value = BWP_slot_dict.get(i)
                            read_SRS_GAP_sheet.cell(nrows+1,slot_num_col+1,subcarrierSpacing_value)
                            
        #UE Index在第一行中
        first_row_list = list[0].split(',')
        #用于记录excel中该ue接入的时间
        for ret in first_row_list:
            #ue_index
            if 'Index' in ret:
                ue_index_value = ret.split(':')[1]          
                read_SRS_GAP_sheet.cell(nrows+1,ue_index_col+1,ue_index_value)
                rrcReconfiguration.append(ue_index_value)
                
                #time(释放时间)
                time_value = first_row_list[0].split(' ')[0]
                if 12 == len(time_value):
                    read_SRS_GAP_sheet.cell(nrows+1,time_col+1,time_value)
                    rrcReconfiguration.append(time_value)
        
                #用于记录excel中该ue接入的时间
                rrcReconfiguration = []       
                num = 1
                #用于记录接入次数
                for i in range(0,len(swich),2):
                    if swich[i] == ue_index_value:
                        rrcReconfiguration.append(swich[i+1])

                ##释放该ue时间rrcReconfiguration = ['18:28:02:283', '18:28:12:741', '18:28:23:224', '18:28:33:763']
                ##上一次接入时间>上一次释放时间>下一次接入时间 
                for j in range(0,len(rrcReconfiguration)-1):
                    #如果第一次释放<第一次接入，则跳过，默认给1
                    if time_value < rrcReconfiguration[0]:
                        break
                    #1）统计每UE发送如下关键标黄字段出现次数N,如N=0，则该UE接入次数记为1
                        #elif time_value < rrcReconfiguration[0]:
                    #2）如N≠0，则判断统计的UE信息落在第n个标黄次数后，则标明该UE为第n+1次接入
                    #如果释放时间大于最大接入时间
                    if len(rrcReconfiguration) > 1:
                        #如果最后一次接入小于释放
                        if rrcReconfiguration[j] < time_value < rrcReconfiguration[j+1]:
                            num = num + j + 1
                            continue
                    if rrcReconfiguration[len(rrcReconfiguration)-1] < time_value:
                        num = 1
                        #为了判断数据中接入次数少于释放次数，则+1
                        if j + 1 == len(rrcReconfiguration)-1:
                            num = num + j + 1 + 1
                        else:
                            num = num + j + 1
                        continue
                        #如果接入卡顿
                    if j > 2:
                        if rrcReconfiguration[j-1] < time_value < rrcReconfiguration[-1]:
                            num = num + j + 1
                            continue
                
                #如果只有一次接入时间
                if len(rrcReconfiguration) == 1:
                    #如果释放时间大于接入时间
                    if time_value > rrcReconfiguration[0]:
                        num = num + 1
                
                read_SRS_GAP_sheet.cell(nrows+1,ue_swich_col+1,str(num))
                        
        #gap相关
        for gap in list:
            if 'measGapConfig setup' in gap:
                gap_index = list.index(gap)
                if 'gapOffset' in list[gap_index+1]:
                    gapOffset_value = list[gap_index+1].split(' ')[1].replace(',','')
                    read_SRS_GAP_sheet.cell(nrows+1,gapOffset_ms_col+1,gapOffset_value)
                if 'mgl' in list[gap_index+2]:
                    mgl_value = list[gap_index+2].split(' ')[1].replace(',','')[2:]
                    read_SRS_GAP_sheet.cell(nrows+1,mgl_ms_col+1,mgl_value)
                if 'mgrp' in list[gap_index+3]:
                    mgrp_value = list[gap_index+3].split(' ')[1].replace(',','')[2:]
                    read_SRS_GAP_sheet.cell(nrows+1,mgrp_ms_col+1,mgrp_value)
                if 'mgta' in list[gap_index+4]:
                    mgta_value = list[gap_index+4].split(' ')[1][2:]
                    read_SRS_GAP_sheet.cell(nrows+1,mgta_col+1,mgta_value)

        #srs-ResourceSetId,最多支持8个,一套为9行
        srs_resourcesetid_list = []
        usage_list = []
        index = 0
        for resourcesetid in list:
            if 'srs-ResourceSetToAddModList' in resourcesetid:
                index = list.index('srs-ResourceSetToAddModList')
        
        if index != 0:                
            for i in range(index,index+72):
                if 'srs-ResourceSetId' in list[i]:
                    srs_resourcesetid = list[i].split(' ')[1].replace(',','')
                    srs_resourcesetid_list.append(srs_resourcesetid)
                if 'usage' in list[i]:
                    usage = list[i].split(' ')[1].replace(',','')
                    usage_list.append(usage)
            if len(srs_resourcesetid_list) > 1: 
                srs_ResourceSetId_value = ''
                for j in range(len(srs_resourcesetid_list)):
                    if j < len(srs_resourcesetid_list)-1:
                        srs_ResourceSetId_value = srs_ResourceSetId_value + srs_resourcesetid_list[j] + '/'
                    else:
                        srs_ResourceSetId_value = srs_ResourceSetId_value + srs_resourcesetid_list[j]
                read_SRS_GAP_sheet.cell(nrows+1,srs_ResourceSetId_col+1,srs_ResourceSetId_value)    
            else:
                read_SRS_GAP_sheet.cell(nrows+1,srs_ResourceSetId_col+1,srs_resourcesetid_list[0])
            if len(usage_list) > 1: 
                usage_value = ''
                for j in range(len(usage_list)):
                    if j < len(usage_list)-1:
                        usage_value = usage_value + usage_list[j] + '/'
                    else:
                        usage_value = usage_value + usage_list[j]
                read_SRS_GAP_sheet.cell(nrows+1,usage_col+1,usage_value)    
            else:
                read_SRS_GAP_sheet.cell(nrows+1,usage_col+1,usage_list[0])
                
        #srs-ResourceId,最多支持8个,一套为23行        
        index = 0
        counts = 1
        for resourceid in list:
            if 'srs-ResourceToAddModList' in resourceid:
                index = list.index('srs-ResourceToAddModList')
                #标记srs-ResourceId及其他参数的个数,避免下方轮询数组下标越界
                counts = str(list).count('nrofSRS-Ports')
        if index != 0:
            for count in range(1,counts+1):
                for i in range(index + (count-1)*23,index+count*23):
                
                    if 'srs-ResourceId' in list[i]:
                        srs_resourceid = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,srs_ResourceId_col+1,srs_resourceid)
        
                    if 'nrofSRS-Ports' in list[i]:
                        nrofSRS_Ports = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,nrofSRS_Ports_col+1,nrofSRS_Ports)
                        
                    if 'transmissionComb' in list[i]:
                        transmissionComb = list[i].split(' ')[1]
                        read_SRS_GAP_sheet.cell(nrows+count,transmissionComb_col+1,transmissionComb)
                        
                    if 'combOffset-n2' in list[i] or 'combOffset-n4' in list[i]:
                        combOffset_n2 = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,combOffset_n2_col+1,combOffset_n2)
                        
                    if 'cyclicShift-n2' in list[i] or 'cyclicShift-n4' in list[i]:
                        cyclicShift_n2 = list[i].split(' ')[1]
                        read_SRS_GAP_sheet.cell(nrows+count,cyclicShift_n2_col+1,cyclicShift_n2)
                        
                    if 'startPosition' in list[i]:
                        startPosition = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,startPosition_col+1,startPosition)
                        
                    if 'nrofSymbols' in list[i]:
                        nrofSymbols = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,nrofSymbols_col+1,nrofSymbols)
                        
                    if 'repetitionFactor' in list[i]:
                        repetitionFactor = list[i].split(' ')[1]
                        read_SRS_GAP_sheet.cell(nrows+count,repetitionFactor_col+1,repetitionFactor)
                        
                    if 'freqDomainPosition' in list[i]:
                        freqDomainPosition = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,freqDomainPosition_col+1,freqDomainPosition)
                        
                    if 'freqDomainShift' in list[i]:
                        freqDomainShift = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,freqDomainShift_col+1,freqDomainShift)
                        
                    if 'c-SRS' in list[i]:
                        c_SRS = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,c_SRS_col+1,c_SRS)
                        
                    if 'b-SRS' in list[i]:
                        b_SRS = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,b_SRS_col+1,b_SRS)
                        
                    if 'b-hop' in list[i]:
                        b_hop = list[i].split(' ')[1].replace(',','')
                        read_SRS_GAP_sheet.cell(nrows+count,b_hop_col+1,b_hop)
                        
                    if 'periodicityAndOffset-p' in list[i]:
                        periodicityAndOffset_p = list[i].split(' ')[1][2:]
                        read_SRS_GAP_sheet.cell(nrows+count,periodicityAndOffset_p_col+1,periodicityAndOffset_p)
                        Offset = list[i].split(' ')[-1]
                        read_SRS_GAP_sheet.cell(nrows+count,Offset_col+1,Offset)
                        
        #合并单元格                
        if counts > 1:
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=ue_index_col+1, end_row=nrows + counts, end_column=ue_index_col+1)
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=time_col+1, end_row=nrows + counts, end_column=time_col+1)
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=bwp_id_col+1, end_row=nrows + counts, end_column=bwp_id_col+1)
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=srs_ResourceSetId_col+1, end_row=nrows + counts, end_column=srs_ResourceSetId_col+1)
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=usage_col+1, end_row=nrows + counts, end_column=usage_col+1)
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=slot_num_col+1, end_row=nrows + counts, end_column=slot_num_col+1)
            read_SRS_GAP_sheet.merge_cells(start_row=nrows+1, start_column=ue_swich_col+1, end_row=nrows + counts, end_column=ue_swich_col+1)
        nrows = nrows + counts
    
    read_workbook.save(path)
    return swich,rrcReconfiguration
    
def calculate_ncs_data(path):
    #打开srs&gap表
    read_workbook = openpyxl.load_workbook(path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']
    #获取SRS表中Ktc的值
    nrows_max = read_SRS_GAP_sheet.max_row
    nrows_min = read_SRS_GAP_sheet.min_row
    cols = read_SRS_GAP_sheet.max_column
    value = []
    for i in range(1,cols + 1):
        cell_value = read_SRS_GAP_sheet.cell(row=1,column=i).value
        value.append(cell_value)
    if 'transmissionComb' in value:
        transmissionComb_col = value.index('transmissionComb')
    if '端口2占用ncs' in value:
        port2_ncs_col = value.index('端口2占用ncs')
    if 'nrofSRS-Ports' in value:
        nrofSRS_Ports_col = value.index('nrofSRS-Ports')
    if 'cyclicShift-n2' in value:
        cyclicShift_n2_col = value.index('cyclicShift-n2')
    #端口2占用ncs
    for j in range(2,nrows_max+1):
        ncs_value = '-'
        #.value = n2 or ports2
        if read_SRS_GAP_sheet.cell(j,transmissionComb_col+1).value != None:
            ktc_value = read_SRS_GAP_sheet.cell(j,transmissionComb_col+1).value[-1]
            cyclicShift_key =  read_SRS_GAP_sheet.cell(j,cyclicShift_n2_col+1).value
            if cyclicShift_key != None:
                port_value = read_SRS_GAP_sheet.cell(j,nrofSRS_Ports_col+1).value[-1]
                if '2' == ktc_value:
                    if port_value != '1':
                        for i in kct2.keys():
                            if i == cyclicShift_key:
                                ncs_value = kct2.get(i)
                                read_SRS_GAP_sheet.cell(j,port2_ncs_col+1,ncs_value)
                    else:
                        read_SRS_GAP_sheet.cell(j,port2_ncs_col+1,ncs_value)
                if '4' == ktc_value:
                    if port_value != '1':
                        for k in kct4.keys():
                            if k == cyclicShift_key:
                                ncs_value = kct4.get(k)
                                read_SRS_GAP_sheet.cell(j,port2_ncs_col+1,ncs_value)
                    else:
                        read_SRS_GAP_sheet.cell(j,port2_ncs_col+1,ncs_value)
    read_workbook.save(path)

def _list_tools(request_transfer):
    #接入去注册消息预处理
    #request_transfer = [接入+去注册]
    removes = []
    for i in range(len(request_transfer)-1):
        list = []
        if request_transfer[i] == '去注册：':
            datas = request_transfer[i+1].replace('\t','').replace('{','').replace('}','').strip(' ')
            for values in datas.split('\n'):
                value = values.strip(' ')
                if value != '':                
                    list.append(value)
            for ret in list:
                #ue_index
                if 'Index' in ret:
                    ue_index_value = ret.split('Index:')[1].split(',')[0].strip()          
                    #time
                    time_value = ret[:12]
                    if 12 == len(time_value):
                        request_transfer[i+1] = ue_index_value + ',' + time_value
        list = []
        if request_transfer[i] == '接入：':
            datas = request_transfer[i+1].replace('\t','').replace('{','').replace('}','').strip(' ')
            for values in datas.split('\n'):
                value = values.strip(' ')
                if value != '':                
                    list.append(value)
            for ret in list:
                #ue_index
                if 'Index' in ret:
                    ue_index_value = ret.split('Index:')[1].split(',')[0].strip()          
                    #time
                    time_value = ret[:12]
                    if 12 == len(time_value):
                        request_transfer[i+1] = ue_index_value + ',' + time_value
        list = []
        if request_transfer[i] == '重配置：':
            datas = request_transfer[i+1].replace('\t','').replace('{','').replace('}','').strip(' ')
            for values in datas.split('\n'):
                value = values.strip(' ')
                if value != '':                
                    list.append(value)
            for ret in list:
                if len(list) > 400:
                    #ue_index
                    if 'Index:' in ret:
                        ue_index_value = ret.split('Index:')[1].split(',')[0].strip()          
                        #time
                        time_value = ret[:12]
                        if 12 == len(time_value):
                            request_transfer[i+1] = ue_index_value + ',' + time_value
                else:
                    request_transfer[i] = '无效消息'
                    request_transfer[i+1] = '无效消息'

    return request_transfer

    
def calculate_srs_conflict_data(path,request_transfer):
    #打开srs&gap表
    read_workbook = openpyxl.load_workbook(path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']
    nrows_max = read_SRS_GAP_sheet.max_row
    nrows_min = read_SRS_GAP_sheet.min_row
    cols = read_SRS_GAP_sheet.max_column
    #读取每一行的K到Y列数据为字符串并添加到list中
    k_y_list = []
    value = []
    for i in range(1,cols + 1):
        cell_value = read_SRS_GAP_sheet.cell(row=1,column=i).value
        value.append(cell_value)
    if 'Ue Index:' in value:
        ue_index_col = value.index('Ue Index:')
    if 'Time' in value:
        time_col = value.index('Time')
    if 'nrofSRS-Ports' in value:
        nrofSRS_Ports_col = value.index('nrofSRS-Ports') + 1
    if 'transmissionComb' in value:
        transmissionComb_col = value.index('transmissionComb') + 1
    if 'combOffset-n2' in value:
        combOffset_n2_col = value.index('combOffset-n2') + 1
    if 'cyclicShift-n2' in value:
        cyclicShift_n2_col = value.index('cyclicShift-n2') + 1
    if '端口2占用ncs' in value:
        port2_ncs_col = value.index('端口2占用ncs') + 1
    if 'startPosition' in value:
        startPosition_col = value.index('startPosition') + 1
    if 'nrofSymbols' in value:
        nrofSymbols_col = value.index('nrofSymbols') + 1
    if 'repetitionFactor' in value:
        repetitionFactor_col = value.index('repetitionFactor') + 1
    if 'freqDomainPosition' in value:
        freqDomainPosition_col = value.index('freqDomainPosition') + 1
    if 'freqDomainShift' in value:
        freqDomainShift_col = value.index('freqDomainShift') + 1
    if 'c-SRS' in value:
        c_SRS_col = value.index('c-SRS') + 1
    if 'b-SRS' in value:
        b_SRS_col = value.index('b-SRS') + 1
    if 'b-hop' in value:
        b_hop_col = value.index('b-hop') + 1
    if 'periodicityAndOffset-p' in value:
        periodicityAndOffset_p_col = value.index('periodicityAndOffset-p') + 1
    if 'Offset' in value:
        Offset_col = value.index('Offset') + 1
    if '各UE资源是否冲突' in value:
        ue_resource_col = value.index('各UE资源是否冲突')
    #循环行
    for nrow in range(2,nrows_max+1):
        k_y_str = ''
        #循环列累加
        for col in range(nrofSRS_Ports_col,Offset_col+1):
            k_y_str = k_y_str + str(read_SRS_GAP_sheet.cell(nrow,col).value)
        #加入列表
        k_y_list.append(k_y_str)
    #判断是否有相同的值
    list_comp = []
    length = len(k_y_list)
    for x in range(length):
        comp1 = ''   
        comp1 = k_y_list[x]
        for y in range(x+1,length):
            comp2 = ''
            comp2 = k_y_list[y]
            if comp1 == comp2:
                list_comp.append(x)
                list_comp.append(y)
            else:
                continue
                
    #接入去注册list整合
    rf = _list_tools(request_transfer)
    dict = {}
    #循环行
    for nrow in range(2,nrows_max+1):
        #存接入消息
        req = []
        tra = []
        if read_SRS_GAP_sheet.cell(nrow,ue_index_col+1).value != None:
            ue_index = (read_SRS_GAP_sheet.cell(nrow,ue_index_col+1).value).strip()
            time_value = read_SRS_GAP_sheet.cell(nrow,time_col+1).value
            for j in range(0,len(rf)-1):
                if rf[j] == '重配置：':
                    if time_value == rf[j+1].split(',')[1]:
                        for k in range(j):
                            if rf[k] == '接入：':
                                ue_request = rf[k+1].split(',')[0]
                                time_request = rf[k+1].split(',')[1]
                                #如果接入的ue_index ==  excel中的
                                if ue_request == ue_index:
                                    #如果接入时间<重配置消息，添加到list中，取最后一个接入时间[-1],为最接近重配置
                                    if time_request < time_value:
                                        req.append(time_request)
                        for l in range(j,len(rf)-1):
                            #收到首个同ue_index的去注册的消息后跳出循环
                            if rf[l] == '去注册：':
                                ue_tranfer = rf[l+1].split(',')[0]
                                time_tranfer = rf[l+1].split(',')[1]
                                if ue_tranfer == ue_index:
                                    if time_tranfer > time_value:
                                        tra.append(time_tranfer)
            #判断如果缺少接入消息或缺少去注册消息时，忽略该次
            if len(req) == 0 or len(tra) == 0:
                continue
            else:
                dict[nrow] = req[-1] + ',' + tra[0]
            
    #list_comp=[0, 225, 1, 106, 1, 202, 1, 310, 1, 426, 2, 117, 2, 214, 2, 314, 2, 427, 3, 228, 3, 439]
    #第0个代表当前下标，第1个代表拥有相同值的下标,row = base_index+1
    #黄色
    fille = PatternFill("solid",fgColor="ffff00")
    list_comp_index = len(list_comp)
    comp1_list = []
    comp2_list = []
    
    for i in range(0,list_comp_index+1,2):
        if i < list_comp_index:
            base_index_value = None
            conflict_index_value = None
        
            #冲突1的下标
            base_index = list_comp[i]
            #冲突2的下标
            conflict_index =  list_comp[i+1]
            #k_y_list[base_index]为第二行数据，row+1为标题行，row+2为第一数据行
            #获取首个下标的ue_index
            if read_SRS_GAP_sheet.cell(base_index+2,ue_index_col+1).value != None:
                #被冲突的ue_index
                ue_index = read_SRS_GAP_sheet.cell(base_index+2,ue_index_col+1).value
                if read_SRS_GAP_sheet.cell(conflict_index+2,ue_index_col+1).value != None:
                    #冲突的ue_index
                    ue_conflict_index = read_SRS_GAP_sheet.cell(conflict_index+2,ue_index_col+1).value
            #获取首个下标的Time
            if read_SRS_GAP_sheet.cell(base_index+2,time_col+1).value != None:
                #被冲突的时间
                base_time = read_SRS_GAP_sheet.cell(base_index+2,time_col+1).value
            if read_SRS_GAP_sheet.cell(conflict_index+2,time_col+1).value != None:
                #冲突的时间
                conflict_time = read_SRS_GAP_sheet.cell(conflict_index+2,time_col+1).value
            #当前表中ue冲突值
            str1 = read_SRS_GAP_sheet.cell(base_index+2,ue_resource_col+1).value
            str2 = read_SRS_GAP_sheet.cell(conflict_index+2,ue_resource_col+1).value
            #与当前循环进行拼接
            if str1 == None or str1 == '':
                base_index_value = (str(conflict_time) + ',' + str(ue_conflict_index)).replace(' ','')
            #如果为空，对比添加，并记录该冲突的接入与去注册时间戳用于后面对比
            else:
                base_index_value = (str(str1) + ';' + '\n' + str(conflict_time) + ',' + str(ue_conflict_index)).replace(' ','')
            if str2 == None or str2 == '':
                conflict_index_value = (str(base_time) + ',' + str(ue_index)).replace(' ','')
            else:
                conflict_index_value = (str(str2) + ';' + '\n' + str(base_time) + ',' + str(ue_index)).replace(' ','')
            #写进excel中
            read_SRS_GAP_sheet.cell(base_index+2,ue_resource_col+1,base_index_value)
            read_SRS_GAP_sheet.cell(conflict_index+2,ue_resource_col+1,conflict_index_value)

    #对其中冲突的execl
    for nrow in range(2,nrows_max+1):
        #如果不存在冲突则跳过
        if read_SRS_GAP_sheet.cell(nrow,ue_resource_col+1).value != None:
            ret = read_SRS_GAP_sheet.cell(nrow,ue_resource_col+1).value
            ret_list = ret.split(',')
            for row in dict.keys():
                if nrow == row :
                    request_time = dict.get(row).split(',')[0]
                    transfer_time = dict.get(row).split(',')[1]
                    #如果只存在一个冲突，判断冲突时间是否在该重配置的接入与去注册中
                    if len(ret_list) == 2:
                        res = ret_list[0]
                        #去dict中找对应的row
                        if request_time < res < transfer_time:
                            break
                        else:
                            read_SRS_GAP_sheet.cell(nrow,ue_resource_col+1,'')
                    if len(ret_list) > 2:
                        ret_list = ret.replace('\n','').split(';')
                        result = ''
                        for res in ret_list:
                            time = res.split(',')[0]
                            #可能会有问题
                            if request_time < time < transfer_time:
                                ue_resource = res + '\n'
                            else:
                                ue_resource = ''
                        read_SRS_GAP_sheet.cell(nrow,ue_resource_col+1,result)
                            
    #进行标黄处理
    for nrow in range(2,nrows_max+1):
        if read_SRS_GAP_sheet.cell(nrow,ue_resource_col+1).value != None:
            #有bug，无法判断为''，用len补漏
            if len(read_SRS_GAP_sheet.cell(nrow,ue_resource_col+1).value) > 12:
                for col in range(1,cols+1):
                    read_SRS_GAP_sheet.cell(nrow,col).fill = fille
    
    read_workbook.save(path)

def calculate_srs_gap(exl_path):
    #打开excel
    read_workbook = openpyxl.load_workbook(exl_path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']
    nrows_max = read_SRS_GAP_sheet.max_row
    nrows_min = read_SRS_GAP_sheet.min_row
    cols = read_SRS_GAP_sheet.max_column
    value = []
    for i in range(1,cols + 1):
        cell_value = read_SRS_GAP_sheet.cell(row=1,column=i).value
        value.append(cell_value)
    #匹配列名
    if 'gapOffset(ms)' in value:
        gapOffset_ms_col = value.index('gapOffset(ms)')
    if 'mgl(ms)' in value:
        mgl_ms_col = value.index('mgl(ms)')
    if 'mgrp(ms)' in value:
        mgrp_ms_col = value.index('mgrp(ms)')
    if 'mgta' in value:
        mgta_col = value.index('mgta')
    if 'periodicityAndOffset-p' in value:
        periodicityAndOffset_p_col = value.index('periodicityAndOffset-p')
    if 'Offset' in value:
        Offset_col = value.index('Offset')
    if '各rssourece     srs落入GAP冲突比例' in value:
        srs_GAP_col = value.index('各rssourece     srs落入GAP冲突比例')
    
    #计算冲突次数
    #gap单位为ms，与srs资源比较时乘以2换算成sl,统一单位
    slot = 2
    #循环行
    for nrow in range(2,nrows_max+1):
        #gapOffset(sl)   18*2
        if read_SRS_GAP_sheet.cell(nrow,gapOffset_ms_col+1).value != None:
            gapOffset = int(read_SRS_GAP_sheet.cell(nrow,gapOffset_ms_col+1).value)*slot
            #mgl(sl)   6*2
            if read_SRS_GAP_sheet.cell(nrow,mgl_ms_col+1).value != None:
                mgl = int(read_SRS_GAP_sheet.cell(nrow,mgl_ms_col+1).value)*slot
            #mgrp(sl)   80*2
            if read_SRS_GAP_sheet.cell(nrow,mgrp_ms_col+1).value != None:
                mgrp = int(read_SRS_GAP_sheet.cell(nrow,mgrp_ms_col+1).value)*slot
            #mgta(sl)   0.5*2
            if read_SRS_GAP_sheet.cell(nrow,mgta_col+1).value != None:
                mgta = int(read_SRS_GAP_sheet.cell(nrow,mgta_col+1).value)*slot
            #periodicityAndOffset(sl)   160
            if read_SRS_GAP_sheet.cell(nrow,periodicityAndOffset_p_col+1).value != None:
                periodicityAndOffset = int(read_SRS_GAP_sheet.cell(nrow,periodicityAndOffset_p_col+1).value)
            #Offset(sl)   47
            if read_SRS_GAP_sheet.cell(nrow,Offset_col+1).value != None:
                Offset = int(read_SRS_GAP_sheet.cell(nrow,Offset_col+1).value)
        else:
            if read_SRS_GAP_sheet.cell(nrow,periodicityAndOffset_p_col+1).value != None:
                gapOffset = 18*slot
                mgl = 6*slot
                mgrp = 80*slot
                mgta = 0.5*slot
                periodicityAndOffset = int(read_SRS_GAP_sheet.cell(nrow,periodicityAndOffset_p_col+1).value)
                if read_SRS_GAP_sheet.cell(nrow,Offset_col+1).value != None:
                    Offset = int(read_SRS_GAP_sheet.cell(nrow,Offset_col+1).value)
                    
        #冲突次数计数
        conflict = 0
        i = 0
        stop = 0
        if gapOffset != None:
            if mgrp >= periodicityAndOffset:
                stop = int(mgrp/periodicityAndOffset - 1)
                for i in range(0,stop+1):
                    if (gapOffset-mgta >= 0) and (gapOffset-mgta+mgl <= mgrp):
                        if (gapOffset-mgta) <= (Offset+i*periodicityAndOffset) <= (gapOffset-mgta+mgl):
                            conflict = conflict + 1
                    else:
                        if ((gapOffset-mgta+mgrp) <= (Offset+i*periodicityAndOffset) <= mgrp) or (0 <= (Offset+i*periodicityAndOffset) <= (gapOffset-mgta+mgrp)):
                            conflict = conflict + 1
            else:
                stop = int(periodicityAndOffset/mgrp - 1)
                for i in range(0,stop+1):
                    if (gapOffset-mgta+i*mgrp) <= Offset <= (gapOffset-mgta+mgl+i*mgrp):
                        conflict = conflict + 1
        else:
            continue
                        
        #SRS冲突比例
        Tmax = 0
        if mgrp >= periodicityAndOffset:
            Tmax = mgrp
        else:
            Tmax = periodicityAndOffset
        n = Tmax/periodicityAndOffset
        conflict_proportion = "%.2f%%"%((conflict/n)*100)
        read_SRS_GAP_sheet.cell(nrow,srs_GAP_col+1,conflict_proportion)
    read_workbook.save(exl_path)
    
    #BWP冲突概率
    #打开excel
    read_workbook = openpyxl.load_workbook(exl_path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']
    nrows_max = read_SRS_GAP_sheet.max_row
    nrows_min = read_SRS_GAP_sheet.min_row
    cols = read_SRS_GAP_sheet.max_column
    value = []
    for i in range(1,cols + 1):
        cell_value = read_SRS_GAP_sheet.cell(row=1,column=i).value
        value.append(cell_value)    
    if 'BWP编号' in value:
        bwp_id_col = value.index('BWP编号')
    if '各rssourece     srs落入GAP冲突比例' in value:
        srs_GAP_col = value.index('各rssourece     srs落入GAP冲突比例')
    if '各BWP上        SRS与GAP冲突概率' in value:
        ue_srs_gap_col = value.index('各BWP上        SRS与GAP冲突概率')
    if 'Time' in value:
        time_col = value.index('Time')
    
    #各资源冲突比例之和/资源数量
    #其中时间戳唯一，可为判断依据
    list_new = []
    bwp_proportion = ''
    total_gap = 0
    merged_cell_list = read_SRS_GAP_sheet.merged_cells.ranges
    if len(merged_cell_list) > 0:
        for merged_cell in merged_cell_list:
            list = []
            rows = re.findall('\d+',str(merged_cell))
            start_row = int(rows[0])
            stop_row = int(rows[1])
            #合并单元格行数的计算
            count = stop_row + 1 - start_row
            list.append(start_row)
            list.append(stop_row)
            list.append(count)
            if len(list_new) > 0:
                if list not in list_new:
                    list_new.append(list)
            else:
                list_new.append(list)
    
    for nrow in range(2,nrows_max+1):
        count = 1
        if read_SRS_GAP_sheet.cell(nrow,column=srs_GAP_col+1).value != None:
            if len(list_new) > 0:
                for ret in list_new:
                    total_gap = 0
                    #ret = [6,8,3]
                    if ret[0] <= nrow <= ret[1]:
                        for i in range(ret[0],ret[1]+1):
                            ue_srs_gap = (float(read_SRS_GAP_sheet.cell(i,column=srs_GAP_col+1).value[:-1]))/100
                            total_gap = total_gap + ue_srs_gap
                        bwp_proportion = "%.2f%%"%((total_gap/(ret[2]))*100)
                        read_SRS_GAP_sheet.cell(ret[0],ue_srs_gap_col+1,bwp_proportion)
                    else:
                        ue_srs_gap = (float(read_SRS_GAP_sheet.cell(nrow,column=srs_GAP_col+1).value[:-1]))/100
                        bwp_proportion = "%.2f%%"%((ue_srs_gap/count)*100)
                        read_SRS_GAP_sheet.cell(nrow,ue_srs_gap_col+1,bwp_proportion)
                    read_SRS_GAP_sheet.merge_cells(start_row=ret[0], start_column=ue_srs_gap_col+1, end_row=ret[1], end_column=ue_srs_gap_col+1)
            else:
                ue_srs_gap = (float(read_SRS_GAP_sheet.cell(nrow,column=srs_GAP_col+1).value[:-1]))/100
                bwp_proportion = "%.2f%%"%((ue_srs_gap/count)*100)
                read_SRS_GAP_sheet.cell(nrow,ue_srs_gap_col+1,bwp_proportion)
                
    read_workbook.save(exl_path)
    
def teardown(exl_path):
    #打开excel
    read_workbook = openpyxl.load_workbook(exl_path,keep_vba=True)
    read_SRS_GAP_sheet = read_workbook['SRSandGAP原始数据']

    #excel加框
    thin = Side(border_style="thin",color="000000")
    nrows_max = read_SRS_GAP_sheet.max_row
    ncols_max = read_SRS_GAP_sheet.max_column

    for i in range(2,nrows_max+1):
        for j in range(1,ncols_max+1):
            #添加边框为黑色
            read_SRS_GAP_sheet.cell(i,j).border=Border(top = thin,right = thin,bottom = thin,left = thin)
            #数据居中
            read_SRS_GAP_sheet.cell(i,j).alignment=Alignment(horizontal='center',vertical='center')
    read_workbook.save(exl_path)
    
#########################################################################################################
#PRACHandPUCCH频域位置填写
#########################################################################################################
#获取需要提取的对应ue1_id
def PRACHandPUCCH_init(path):
    #打开srs&gap表
    read_workbook = openpyxl.load_workbook(path,keep_vba=True)
    read_PRACHandPUCCH_sheet = read_workbook['PRACHandPUCCH频域位置填写']
    nrows_max = read_PRACHandPUCCH_sheet.max_row
    nrows_min = read_PRACHandPUCCH_sheet.min_row
    cols = read_PRACHandPUCCH_sheet.max_column
    for row in range(1,nrows_max + 1):
        for col in range(1,cols + 1):
            #由于合并单元格原因，49行为标题，需要+2获取到值
            if read_PRACHandPUCCH_sheet.cell(row,col).value == '需要提取的对应ue1_id':
                ue_id1 = read_PRACHandPUCCH_sheet.cell(row+2,col).value
            if read_PRACHandPUCCH_sheet.cell(row,col).value == '需要提取的对应ue2_id':
                ue_id2 = read_PRACHandPUCCH_sheet.cell(row+2,col).value
    if ue_id1 == None and ue_id2 == None:
        raise "请联系自动化组，逻辑已更改，需要适配"
    read_workbook.save(path)
    return ue_id1,ue_id2

#获取自动录入参数并填写
def set_PRACHandPUCCH_auto_data(invalid_datas,auto_datas,path,ue_id1,ue_id2):
    for invalid_data in invalid_datas:
        list = []
        #对每一个数据去除特殊字符，并按照回车切分
        datas = invalid_data.replace('\t','').replace('{','').replace('}','').strip(' ')
        #对每行数据进行去除前后的空格
        for data in datas.split('\n'):
            value = data.strip(' ')
            if value != '':                
                list.append(value)
        
        #初始值
        freqBandIndicatorNR = ''
        carrierBandwidth = ''
        locationAndBandwidth = ''
        subcarrierSpacing = ''
        prach_ConfigurationIndex = ''
        msg1_FDM = ''
        msg1_FrequencyStart = ''
        pucch_ResourceCommon = ''
        for i in list:     
            if 'freqBandIndicatorNR' in i:
                freqBandIndicatorNR = i.split(' ')[1]
            if 'carrierBandwidth' in i:
                carrierBandwidth = i.split(' ')[1]
            if 'locationAndBandwidth' in i:
                locationAndBandwidth = i.split(' ')[1][:-1]
            if 'subcarrierSpacing' in i:
                subcarrierSpacing = i.split(' ')[1][3:]
            if 'prach-ConfigurationIndex' in i:
                prach_ConfigurationIndex = i.split(' ')[1][:-1]
            if 'msg1-FDM' in i:
                msg1_FDM_str = i.split(' ')[1][:-1]
                msg1_FDM = _num_tools(msg1_FDM_str)
            if 'msg1-FrequencyStar' in i:
                msg1_FrequencyStart = i.split(' ')[1][:-1]
            if 'pucch-ResourceCommon' in i:
                pucch_ResourceCommon = i.split(' ')[1][:-1]    
    
    for auto_data in auto_datas:
        list = []
        datas = auto_data.replace('\t','').replace('{','').replace('}','').strip(' ')
        #对每行数据进行去除前后的空格
        for data in datas.split('\n'):
            value = data.strip(' ')
            if value != '':                
                list.append(value)
        
        #UE Index在第一行中
        first_row_list = list[0].split(',')
        for ret in first_row_list:
            #ue_index
            if 'Index' in ret:
                ue_index_value = (ret.split(':')[1]).strip()
                
        #初始值
        locationAndBandwidth_bwp1 = ''
        subcarrierSpacing_bwp1 = ''
        prach_ConfigurationIndex_bwp1 = ''
        msg1_FDM_bwp1 = ''
        msg1_FrequencyStart_bwp1 = ''
        locationAndBandwidth_bwp2 = ''
        subcarrierSpacing_bwp2 = ''
        prach_ConfigurationIndex_bwp2 = ''
        msg1_FDM_bwp2 = ''
        msg1_FrequencyStart_bwp2 = ''
        list1 = []
        list2 = []
        if str(ue_id1) == ue_index_value:
            for i in range(len(list)):
                if 'uplinkBWP-ToAddModList' in list[i]:
                    if 'bwp-Id 1' in list[i+1]:
                        locationAndBandwidth_bwp1 = list[i+4].split(' ')[1]
                        subcarrierSpacing_bwp1 = list[i+5].split(' ')[1][3:]
                        prach_ConfigurationIndex_bwp1 = list[i+9].split(' ')[1][:-1]
                        msg1_FDM_bwp1_str = list[i+10].split(' ')[1][:-1]
                        msg1_FDM_bwp1 = _num_tools(msg1_FDM_bwp1_str)
                        msg1_FrequencyStart_bwp1 = list[i+11].split(' ')[1][:-1]
                        list1.append(msg1_FDM_bwp1)
                    if 'bwp-Id 2' in list[i+1]:
                        locationAndBandwidth_bwp2 = list[i+4].split(' ')[1]
                        subcarrierSpacing_bwp2 = list[i+5].split(' ')[1][3:]
                        prach_ConfigurationIndex_bwp2 = list[i+9].split(' ')[1][:-1]
                        msg1_FDM_bwp2_str = list[i+10].split(' ')[1][:-1]
                        msg1_FDM_bwp2 = _num_tools(msg1_FDM_bwp2_str)
                        msg1_FrequencyStart_bwp2 = list[i+11].split(' ')[1][:-1]
                        list2.append(msg1_FDM_bwp2)
        #判断当找到bwp1或2中参数时，跳出循环，填写进excel
        if len(list1)>0 or len(list2)>0:
            break
                
    #打开srs&gap表
    read_workbook = openpyxl.load_workbook(path,keep_vba=True)
    read_PRACHandPUCCH_sheet = read_workbook['PRACHandPUCCH频域位置填写']
    #获取SRS表中Ktc的值
    nrows_max = read_PRACHandPUCCH_sheet.max_row
    nrows_min = read_PRACHandPUCCH_sheet.min_row
    list = []
    #用于储存多个相同行的行号
    subcarrierSpacing_list = []
    prach_ConfigurationIndex_list = []
    msg1_FDM_list = []
    msg1_FrequencyStart_list = []
    for nrow in range(nrows_min,nrows_max):
        if None != read_PRACHandPUCCH_sheet.cell(nrow,1).value:
            if 'freqBandIndicatorNR' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                read_PRACHandPUCCH_sheet.cell(nrow,2,freqBandIndicatorNR)
            if 'carrierBandwidth' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                read_PRACHandPUCCH_sheet.cell(nrow,2,carrierBandwidth)
            if 'initialUplinkBWP(BWP0)' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                read_PRACHandPUCCH_sheet.cell(nrow,2,locationAndBandwidth)
            if 'initialUplinkBWP(BWP1)' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                read_PRACHandPUCCH_sheet.cell(nrow,2,locationAndBandwidth_bwp1)
            if 'initialUplinkBWP(BWP2)' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                read_PRACHandPUCCH_sheet.cell(nrow,2,locationAndBandwidth_bwp2)
            if 'subcarrierSpacing' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                subcarrierSpacing_list.append(nrow)
            if 'prach-ConfigurationIndex' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                prach_ConfigurationIndex_list.append(nrow)
            if 'msg1-FDM' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                msg1_FDM_list.append(nrow)    
            if 'msg1-FrequencyStart' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                msg1_FrequencyStart_list.append(nrow)
            if 'pucch-ResourceCommon' in read_PRACHandPUCCH_sheet.cell(nrow,1).value:
                read_PRACHandPUCCH_sheet.cell(nrow,2,pucch_ResourceCommon)

    read_PRACHandPUCCH_sheet.cell(subcarrierSpacing_list[0],2,subcarrierSpacing)
    read_PRACHandPUCCH_sheet.cell(subcarrierSpacing_list[1],2,subcarrierSpacing_bwp1)
    read_PRACHandPUCCH_sheet.cell(subcarrierSpacing_list[2],2,subcarrierSpacing_bwp2)
    read_PRACHandPUCCH_sheet.cell(prach_ConfigurationIndex_list[0],2,prach_ConfigurationIndex)
    read_PRACHandPUCCH_sheet.cell(prach_ConfigurationIndex_list[1],2,prach_ConfigurationIndex_bwp1)
    read_PRACHandPUCCH_sheet.cell(prach_ConfigurationIndex_list[2],2,prach_ConfigurationIndex_bwp2)
    read_PRACHandPUCCH_sheet.cell(msg1_FDM_list[1],2,msg1_FDM)
    read_PRACHandPUCCH_sheet.cell(msg1_FDM_list[2],2,msg1_FDM_bwp1)
    read_PRACHandPUCCH_sheet.cell(msg1_FDM_list[3],2,msg1_FDM_bwp2)
    read_PRACHandPUCCH_sheet.cell(msg1_FrequencyStart_list[1],2,msg1_FrequencyStart)
    read_PRACHandPUCCH_sheet.cell(msg1_FrequencyStart_list[2],2,msg1_FrequencyStart_bwp1)
    read_PRACHandPUCCH_sheet.cell(msg1_FrequencyStart_list[3],2,msg1_FrequencyStart_bwp2)
    
    read_workbook.save(path)
    
def _num_tools(str):
    num = ''
    dict = {'one':'1','two':'2','three':'3','four':'4'}
    for i in dict.keys():
        if i == str:
            num = dict.get(i)
    return num
    
def get_PRACHandPUCCH_resoure(auto_datas,exl_path,ue_id1,ue_id2):
    #清空excel
    #打开srs&gap表
    read_workbook = openpyxl.load_workbook(exl_path,keep_vba=True)
    read_PRACHandPUCCH_sheet = read_workbook['PRACHandPUCCH频域位置填写']
    nrows_max = read_PRACHandPUCCH_sheet.max_row
    nrows_min = read_PRACHandPUCCH_sheet.min_row
    cols = read_PRACHandPUCCH_sheet.max_column
    resources_col = []
    startingPRB_col = []
    for row in range(1,nrows_max + 1):
        for col in range(1,cols + 1):
            if read_PRACHandPUCCH_sheet.cell(row,col).value == 'startingPRB':
                min_row = row
                startingPRB_col.append(col)
            if read_PRACHandPUCCH_sheet.cell(row,col).value == '各resoure对应的资源属性':
                resources_col.append(col)
                max_col = col
    for i in range(min_row+1,nrows_max+1):
        for j in range(1,max_col+1):
            read_PRACHandPUCCH_sheet.cell(i,j,'')
    read_workbook.save(exl_path)
    
    sr1_list = []
    set0_list1 = []
    set1_list1 = []
    ue1 = []
    ue2 = []
    for auto_data in auto_datas:
        list = []
        datas = auto_data.replace('\t','').replace('{','').replace('}','').strip(' ')
        #对每行数据进行去除前后的空格
        for data in datas.split('\n'):
            value = data.strip(' ')
            if value != '':                
                list.append(value)
        
        #UE Index在第一行中
        first_row_list = list[0].split(',')
        for ret in first_row_list:
            #ue_index
            if 'Index' in ret:
                ue_index_value = (ret.split(':')[1]).strip()
        
        #对其中数据进一步处理
        #1.根据ue_id1拆分有用数据
        ue1_list = []
        if str(ue_id1) == ue_index_value:
            for i in range(len(list)):
                if 'uplinkBWP-ToAddModList' in list[i]:
                    row_start = i
                if 'firstActiveUplinkBWP-Id' in list[i]:
                    row_stop = i
                    ue1_list = list[row_start:row_stop]
        if len(ue1_list) > 0:
            ue1.append(ue1_list)

        #1.根据ue_id2拆分有用数据
        ue2_list = []
        if str(ue_id2) == ue_index_value:
            for i in range(len(list)):
                if 'uplinkBWP-ToAddModList' in list[i]:
                    row_start = i
                if 'firstActiveUplinkBWP-Id' in list[i]:
                    row_stop = i
                    ue2_list = list[row_start:row_stop]
        if len(ue2_list) > 0:
            ue2.append(ue2_list)
                    
        #2根据BWP区分list，当前不存在BWP2，未知结构，暂不写
        #若基站配置相同，则不同ue_index的pucchresourceid相同
        sr1_list = []
        if str(ue_id1) == ue_index_value:
            for i in range(len(ue1_list)-2):
                if 'schedulingRequestID' in ue1_list[i]:
                    if 'resource' in ue1_list[i+2]:
                        sr1_list.append(ue1_list[i+2].split(' ')[1])
                        
        #set01
        set0_list1 = []
        set1_list1 = []
        if 'pucch-ResourceSetId 0,' in ue1_list:
            if 'pucch-ResourceSetId 1,' in ue1_list:
                for i in range(len(ue1_list)):
                    if 'pucch-ResourceSetId 0' in ue1_list[i]:
                        row_start = i
                    if 'pucch-ResourceSetId 1' in ue1_list[i]:
                        row_stop = i
                        data = get_PCO_RRC_NR_DL_CCCH_Message(row_start,row_stop,ue1_list)
                        if '\n' in data:
                            datas = data.replace(',','').split('\n')
                        for value in datas:
                            try:
                                #如果无法转换成int类型，则跳过
                                d = int(value)
                                set0_list1.append(value)
                            except:
                                continue
                    if 'pucch-ResourceSetId 1,' in ue1_list[i]:
                        row_start1 = i
                    if 'resourceToAddModList' in ue1_list[i]:
                        row_stop1 = i
                        data = get_PCO_RRC_NR_DL_CCCH_Message(row_start1,row_stop1,ue1_list)
                        if '\n' in data:
                            datas = data.replace(',','').split('\n')
                        for value in datas:
                            try:
                                #如果无法转换成int类型，则跳过
                                d = int(value)
                                set1_list1.append(value)
                            except:
                                continue
            else:
                for i in range(len(ue1_list)):
                    if 'pucch-ResourceSetId 0,' in ue1_list[i]:
                        row_start = i
                    if 'resourceToAddModList' in ue1_list[i]:
                        row_stop = i
                        data = get_PCO_RRC_NR_DL_CCCH_Message(row_start,row_stop,ue1_list)
                        if '\n' in data:
                            datas = data.replace(',','').split('\n')
                        for value in datas:
                            try:
                                #如果无法转换成int类型，则跳过
                                d = int(value)
                                set0_list1.append(value)
                            except:
                                continue
        else:
            if 'pucch-ResourceSetId 1,' in ue1_list:
                for i in range(len(ue1_list)):
                    if 'pucch-ResourceSetId 1,' in ue1_list[i]:
                        row_start1 = i
                    if 'resourceToAddModList' in ue1_list[i]:
                        row_stop1 = i
                        data = get_PCO_RRC_NR_DL_CCCH_Message(row_start1,row_stop1,ue1_list)
                        if '\n' in data:
                            datas = data.replace(',','').split('\n')
                        for value in datas:
                            try:
                                #如果无法转换成int类型，则跳过
                                d = int(value)
                                set1_list1.append(value)
                            except:
                                continue
    
        #区别与反复接入,只获取首次有效数据
        if len(ue1) > 0 and len(ue2) > 0:
            if len(sr1_list)>0 or len(set0_list1)>0 or len(set1_list1)>0:
                break
    #先填入资源列
    for i in range(len(sr1_list)):
        for row in range(min_row+1,nrows_max+1):
            read_PRACHandPUCCH_sheet.cell(row+i,1,'pucch-ResourceId ' + sr1_list[i])
            read_PRACHandPUCCH_sheet.cell(row+i,resources_col[0],'sr')
            read_PRACHandPUCCH_sheet.cell(row+i,resources_col[1],'sr')
            break
    for j in range(len(set0_list1)):
        for row in range(min_row + 1 +len(sr1_list),nrows_max+1):
            read_PRACHandPUCCH_sheet.cell(row+j,1,'pucch-ResourceId ' + set0_list1[j])
            read_PRACHandPUCCH_sheet.cell(row+j,resources_col[0],'set0')
            read_PRACHandPUCCH_sheet.cell(row+j,resources_col[1],'set0')
            break
    for k in range(len(set1_list1)):
        for row in range(min_row + 1 + len(sr1_list) + len(set0_list1),nrows_max+1):
            read_PRACHandPUCCH_sheet.cell(row+k,1,'pucch-ResourceId ' + set1_list1[k])
            read_PRACHandPUCCH_sheet.cell(row+k,resources_col[0],'set1')
            read_PRACHandPUCCH_sheet.cell(row+k,resources_col[1],'set1')
            break
    read_workbook.save(exl_path)
    
    #获取对应resourcesId对应的startingPRB
    for row in range(min_row+1,nrows_max+1):
        if read_PRACHandPUCCH_sheet.cell(row,1).value != None:
            resourcesId =  read_PRACHandPUCCH_sheet.cell(row,1).value
            list = ue1[0]
            for i in range(len(list)-1):
                if (len(resourcesId) > 0):
                    if resourcesId in list[i]:
                        startingPRB = list[i+1].replace(',','').split(' ')[1]
                        read_PRACHandPUCCH_sheet.cell(row,startingPRB_col[0],startingPRB)
                else: continue
        else: break
        
    for row in range(min_row+1,nrows_max+1):
        if read_PRACHandPUCCH_sheet.cell(row,1).value != None:
            resourcesId =  read_PRACHandPUCCH_sheet.cell(row,1).value
            list = ue2[0]
            for i in range(len(list)-1):
                if (len(resourcesId) > 0):
                    if resourcesId in list[i]:
                        startingPRB = list[i+1].replace(',','').split(' ')[1]
                        read_PRACHandPUCCH_sheet.cell(row,startingPRB_col[1],startingPRB)
                else: continue
        else: break
        
    read_workbook.save(exl_path)

def get_PRACHandPUCCH_csi(data_list,exl_path,ue_id1,ue_id2):
    #清空excel
    #打开srs&gap表
    read_workbook = openpyxl.load_workbook(exl_path,keep_vba=True)
    read_PRACHandPUCCH_sheet = read_workbook['PRACHandPUCCH频域位置填写']
    nrows_max = read_PRACHandPUCCH_sheet.max_row
    nrows_min = read_PRACHandPUCCH_sheet.min_row
    cols = read_PRACHandPUCCH_sheet.max_column
    resources_col = []
    startingPRB_col = []
    for row in range(1,nrows_max + 1):
        for col in range(1,cols + 1):
            if read_PRACHandPUCCH_sheet.cell(row,col).value == 'startingPRB':
                startingPRB_col.append(col)
            if read_PRACHandPUCCH_sheet.cell(row,col).value == '各resoure对应的资源属性':
                resources_col.append(col)
        if read_PRACHandPUCCH_sheet.cell(row,1).value != None:
            csi_row = row + 1

    ResourceId = []
    startingPRB1 = []
    startingPRB2 = []
    for valid_data in data_list:
        list = []
        ue_index_value = ''
        #对每一个数据去除特殊字符，并按照回车切分
        datas = valid_data.replace('\t','').replace('{','').replace('}','').strip(' ')
        #对每行数据进行去除前后的空格
        for data in datas.split('\n'):
            value = data.strip(' ')
            if value != '':                
                list.append(value)
        
        #判断消息长度，是否有必要存入excel中
        if len(list) < 400:
            continue
            
        #UE Index在第一行中
        first_row_list = list[0].split(',')
        #用于记录excel中该ue接入的时间
        for ret in first_row_list:
            #ue_index
            if 'Index' in ret:
                ue_index_value = ret.split(':')[1].replace(' ','')          

        pucch_ResourceId_list = []
        startingPRB_list1 = []
        startingPRB_list2 = []
        for i in range(len(list)):
            if 'pucch-ResourceId' in list[i]:
                if ue_index_value == str(ue_id1):
                    pucch_ResourceId = list[i].replace(',','').split(' ')[1]
                    pucch_ResourceId_list.append('pucch-ResourceId ' + pucch_ResourceId)
                    if 'startingPRB' in list[i+1]:
                        startingPRB = list[i+1].replace(',','').split(' ')[1]
                        startingPRB_list1.append(startingPRB)
                if ue_index_value == str(ue_id2):
                    if 'startingPRB' in list[i+1]:
                        startingPRB = list[i+1].replace(',','').split(' ')[1]
                        startingPRB_list2.append(startingPRB)
        if len(pucch_ResourceId_list) > 0:
            ResourceId.append(pucch_ResourceId_list)
        if len(startingPRB_list1) > 0:
            startingPRB1.append(startingPRB_list1)
        if len(startingPRB_list2) > 0:
            startingPRB2.append(startingPRB_list2)
        
    for i in range(len(ResourceId[0])):
        read_PRACHandPUCCH_sheet.cell(csi_row,1,ResourceId[0][i])
        read_PRACHandPUCCH_sheet.cell(csi_row,startingPRB_col[0],startingPRB1[0][i])
        read_PRACHandPUCCH_sheet.cell(csi_row,startingPRB_col[1],startingPRB2[0][i])
        read_PRACHandPUCCH_sheet.cell(csi_row,resources_col[0],'csi')
        read_PRACHandPUCCH_sheet.cell(csi_row,resources_col[1],'csi')
        csi_row = csi_row + 1

                
    read_workbook.save(exl_path)

    
print('工具初始化...')
exl_path,log_path = init()
print('获取log信息..')
valid_data,swich_data,request_transfer,invalid_data,auto_data = get_data(log_path)
print('SRS资源分配结果提取中...')
ue_swich,ue_rrcReconfiguration = set_GAP_AND_SRS(exl_path,valid_data,swich_data)
print('计算相关数据中...')
calculate_ncs_data(exl_path)
calculate_srs_conflict_data(exl_path,request_transfer)
#计算各rssourece srs落入GAP冲突比例 与 各BWP上SRS与GAP冲突概率
calculate_srs_gap(exl_path)
print('已完成SRS资源分配结果...')
#后处理单元格
teardown(exl_path)
print('美化SRS资源分配完成...')

print('PRACHandPUCCH频域位置填写初始化...')
print('获取用户填写的ue_index...')
ue_id1,ue_id2 = PRACHandPUCCH_init(exl_path)
print('填写自动录入参数...')
set_PRACHandPUCCH_auto_data(invalid_data,auto_data,exl_path,ue_id1,ue_id2)
print('获取pucch相关资源...')
get_PRACHandPUCCH_resoure(auto_data,exl_path,ue_id1,ue_id2)
get_PRACHandPUCCH_csi(valid_data,exl_path,ue_id1,ue_id2)


print('已完成，请查看...')


